# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from Netbox.DUO_test.datas.MyLog import MyLogs
import os


class ReadDatas:

    def __init__(self, filename):
        self.log = MyLogs()
        if not os.path.exists(filename):
            self.log.error("找不到文件")
        else:
            self.filename = filename

    def read_datas(self, sheetname):
        try:
            wb = load_workbook(self.filename)
        except Exception as e:
            self.log.error("打开文件异常{}".format(e))
        else:
            if sheetname not in wb.sheetnames:
                self.log.error("找不到表单")
            else:
                sheet = wb[sheetname]
                test_name = []
                for column in range(1, sheet.max_column):
                    test_name.append(sheet.cell(1, column).value)
                testdata = []
                for i in range(2, sheet.max_row + 1):
                    testdict = {}
                    for j in range(1, sheet.max_column):
                        testdict[test_name[j - 1]] = sheet.cell(i, j).value
                    testdata.append(testdict)
                return testdata

    def write_back(self, sheetname, row, wifi, intensity, gps, rtc, sim, fv, sd, iden, rst):

        try:
            wb = load_workbook(self.filename)
        except Exception as e:
            self.log.error("打开文件异常{}".format(e))
        else:
            if sheetname not in wb.sheetnames:
                self.log.error("找不到文件")
            else:
                sheet = wb[sheetname]
                # sheet.cell(row, vertical).value = result
                sheet.cell(row, 3).value = wifi
                sheet.cell(row, 4).value = intensity
                sheet.cell(row, 5).value = gps
                sheet.cell(row, 6).value = rtc
                sheet.cell(row, 7).value = sim
                sheet.cell(row, 8).value = fv
                sheet.cell(row, 9).value = sd
                sheet.cell(row, 10).value = iden
                sheet.cell(row, 11).value = rst
                wb.save(self.filename)

    def write_wifi(self, sheetname, row, result):

        try:
            wb = load_workbook(self.filename)
        except Exception as e:
            self.log.error("打开文件异常{}".format(e))
        else:
            if sheetname not in wb.sheetnames:
                self.log.error("找不到文件")
            else:
                sheet = wb[sheetname]
                # sheet.cell(row, vertical).value = result
                sheet.cell(row, 3).value = result
                wb.save(self.filename)
