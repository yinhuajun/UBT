# -*- coding:utf-8 -*-
# @Time     :2019/4/8 10:07
# @Author   :YIN
# @Email    :649626809@qq.com
# @File     :read_datas.py
# @software :PyCharm
from openpyxl import load_workbook
from BusProject.Linebus_Platform_API.common.MyLog import MyLogs
import os


class ReadDatas:

    def __init__(self, filename):
        self.log = MyLogs()
        if not os.path.exists(filename):
            self.log.error("找不到文件")
        else:
            self.filename = filename

    def read_datas(self, sheetname):
        try:
            wb = load_workbook(self.filename)
        except Exception as e:
            self.log.error("打开文件异常{}".format(e))
        else:
            if sheetname not in wb.sheetnames:
                self.log.error("找不到表单")
            else:
                sheet = wb[sheetname]
                test_name = []
                for column in range(1, sheet.max_column):
                    test_name.append(sheet.cell(1, column).value)
                testdata = []
                for i in range(2, sheet.max_row + 1):
                    testdict = {}
                    for j in range(1, 9):
                        testdict[test_name[j - 1]] = sheet.cell(i, j).value
                    testdata.append(testdict)
                return testdata

    def write_back(self, sheetname, row, result, passed):
        try:
            wb = load_workbook(self.filename)
        except Exception as e:
            self.log.error("打开文件异常{}".format(e))
        else:
            if sheetname not in wb.sheetnames:
                self.log.error("找不到文件")
            else:
                sheet = wb[sheetname]
                sheet.cell(row, 9).value = result
                sheet.cell(row, 10).value = passed
                wb.save(self.filename)

